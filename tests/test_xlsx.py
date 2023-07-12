from openpyxl import load_workbook
import os


def test_xlsx_file():
    xlsx_file = os.path.abspath(os.path.join(os.path.dirname(__file__), '../resources/file_example_XLSX_50.xlsx'))
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    print(sheet.cell(row=5, column=5).value)
    assert sheet.cell(row=5, column=5).value == 'United States'
